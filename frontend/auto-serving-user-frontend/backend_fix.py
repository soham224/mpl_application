import io
import logging
from typing import List, Any
from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ... existing imports and dependencies ...

@router.post("/get_result_for_result_excel")
@rate_limiter()
def get_result_for_result_excel(
    request: Request,
    id_list: List[str],
    db: Session = Depends(deps.get_db),
    current_user: models.User = Depends(deps.get_current_active_user),
) -> Any:
    logging.info("START: get_result_for_result_excel")
    
    if len(id_list) == 0:
        logging.error("No IDs provided")
        raise HTTPException(status_code=404, detail="Data Not Found")
    
    try:
        logging.info("Fetching cameras")
        deployment_camera_list = crud.deployment_camera.get_all_camera_by_company_id(
            db, company_id=current_user.company_id
        )
        logging.info("Fetched cameras: %s", deployment_camera_list)
        
        camera_name_dict = {
            camera_detail.id: [
                camera_detail.camera_name,
                camera_detail.location_details.location_name,
            ]
            for camera_detail in deployment_camera_list
        }
        logging.info("Camera name dict: %s", camera_name_dict)
        
        logging.info("Fetching results by ids")
        data_list = get_result_by_ides(id_list)
        logging.info("Fetched data_list: %s", data_list)
        
        if not data_list:
            logging.error("No data found for the provided IDs")
            raise HTTPException(status_code=404, detail="No data found for the provided IDs")
        
        logging.info("Creating excel file")
        output_file = create_excel_file(
            data_list=data_list, camera_name_dict=camera_name_dict
        )
        logging.info("Created output_file: %s", output_file)
        
        # Ensure output_file is a file-like object
        if not hasattr(output_file, "read"):
            logging.error("create_excel_file did not return a file-like object")
            raise HTTPException(status_code=500, detail="Excel file generation failed")
        
        # Reset file pointer to beginning
        output_file.seek(0)
        
        logging.info("Returning StreamingResponse")
        return StreamingResponse(
            output_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=job_results.xlsx",
                "Cache-Control": "no-cache",
                "Pragma": "no-cache",
            },
        )
        
    except Exception as e:
        logging.error("Error in get_result_for_result_excel: %s", str(e))
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

def create_excel_file(data_list, camera_name_dict):
    """
    Create Excel file from data list
    """
    try:
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Results"
        
        # Define headers
        headers = [
            "ID", "Camera Name", "Location", "Created Date", "Labels", "Count"
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Write data
        for row, data in enumerate(data_list, 2):
            camera_id = data.get("camera_id")
            camera_info = camera_name_dict.get(camera_id, ["Unknown", "Unknown"])
            
            # Handle MongoDB ObjectId conversion
            id_value = data.get("_id", "")
            if isinstance(id_value, dict) and "$oid" in id_value:
                id_value = id_value["$oid"]
            elif hasattr(id_value, '__str__'):
                id_value = str(id_value)
            
            # Handle created_date conversion
            created_date = data.get("created_date", "")
            if isinstance(created_date, dict) and "$date" in created_date:
                created_date = created_date["$date"]
            elif hasattr(created_date, '__str__'):
                created_date = str(created_date)
            
            # Handle counts/labels
            counts = data.get("counts", {})
            labels = ", ".join(counts.keys()) if counts else ""
            
            # Handle detection count
            result = data.get("result", {})
            detection = result.get("detection", []) if result else []
            count = len(detection) if detection else 0
            
            ws.cell(row=row, column=1, value=id_value)
            ws.cell(row=row, column=2, value=camera_info[0] if camera_info else "Unknown")
            ws.cell(row=row, column=3, value=camera_info[1] if camera_info else "Unknown")
            ws.cell(row=row, column=4, value=created_date)
            ws.cell(row=row, column=5, value=labels)
            ws.cell(row=row, column=6, value=count)
        
        # Create file-like object in memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
        
    except Exception as e:
        logging.error("Error creating Excel file: %s", str(e))
        raise Exception(f"Failed to create Excel file: {str(e)}") 